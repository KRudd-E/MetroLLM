# IMPORTS
import yaml
import os
import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv
from ollama import Client
from openai import OpenAI
from tqdm import tqdm
from time import gmtime, strftime


class Task:
    def __init__(self):
        self.config = self.get_config()

    def run(self):
        self.user_query()
        df = self.get_df()
        df_up = self.new_tasks(df)
        self.save_to_csv(df_up)


    def get_config(self):
        with open(os.path.join(os.getcwd() + "/_misc/updatedTask/config.yaml"), "r") as file:
            config = yaml.safe_load(file)
        return config    
    
    def user_query(config):
        while True:
            user_input = input("This will generate new tasks for the excel document. Do you wish to proceed (yes/no): ").strip().lower()
            if user_input not in ['yes', 'y', 'no', 'n']:
                print("Invalid input. Please enter 'yes' or 'no'.")
            elif user_input in ['no', 'n']:
                print("Exiting the pipeline.")
                exit(0)
            else:
                print("Proceeding.")
                break

    def check_for_task_cols(self):
        #! Not in use, as writing directly to Excel file causes issues with comments.
        """Checks if the Excel file has the required columns for new tasks."""
        wb = load_workbook(os.path.join(os.getcwd() + self.config['file_dir']))
        ws = wb.active
        if str(ws[1][4].value) != 'Task' or str(ws[1][5].value) != 'Task':
            raise ValueError(f"Incorrect column layout in Excel file. Ensure E1 & F1 are 'Task.\n Currently: {ws[1][4].value} and {ws[1][5].value}.")

    
    def get_df(self):
        """Extracts case study information from the Excel file.

        Returns:
            df:  DataFrame containing case study information
        """
        wb = load_workbook(os.path.join(os.getcwd() + self.config['file_dir']))
        ws = wb.active
        df = pd.DataFrame(columns=['Name', 'origTask', 'newTask1','newTask2',  'subfolder_path', 'subfolder_name', 'file_name'], index=range(1, ws.max_row))
        
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
            cell = row[0]
            try:
                pdf_path = cell.hyperlink.target.replace('%20', ' ')
                file_name = (os.path.basename(pdf_path)[:-4] + '.txt')
                df.at[cell.row - 1, 'Name'] = cell.value
                subfolder_name = os.path.basename(os.path.dirname(pdf_path))
                subfolder_path = os.path.join(os.getcwd() + self.config['txt_grandparent_dir'] + subfolder_name)
                if not os.path.exists(subfolder_path):
                    tqdm.write(f"Missing subfolder: {subfolder_path}")
                    continue
            except AttributeError:
                tqdm.write(f"{cell.coordinate} - No hyperlink in cell ")
                continue
            
            
            df.at[cell.row - 1, 'subfolder_name'] = subfolder_name
            df.at[cell.row - 1, 'subfolder_path'] = subfolder_path
            df.at[cell.row - 1, 'file_name'] = file_name
        
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
            cell = row[0]
            if cell.value is not None:
                df.at[cell.row - 1, 'origTask'] = cell.value
        return df
    
    
    def new_tasks(self, df):
        for index, row in tqdm(df.iterrows(), total=len(df), desc="Generating new tasks"):
            if row['origTask'] == 'Diverse':
                df.at[index, 'newTask1'] = 'Diverse'
                tqdm.write(f"Skipping row {index} - Diverse task")
                continue
            if pd.isna(row['subfolder_path']) or pd.isna(row['file_name']):
                continue
            try:
                with open(os.path.join(row['subfolder_path'] + '/' + row['file_name']), 'r', encoding='utf-8') as file:
                    txt = file.read().replace('-\n','').replace('\n', ' ').replace('\r', ' ')
                    task_list = self.tasks_string(self.config['updated_tasks'])
                    input =  self.config['prompt'].format(origTask=df.at[index,'origTask'],txt=txt, task_list=task_list)
                    i, output_tasks = 1, []
                    while not output_tasks and i < 6:
                        output = self.ai_chat(self.config['model_source'], self.config['model'], input)
                        output_tasks = self.get_tasks(output, self.config['updated_tasks'])
                        i += 1
                    if output_tasks:
                        df.at[index, 'newTask1'] = output_tasks[0] if len(output_tasks) > 0 else None
                        df.at[index, 'newTask2'] = output_tasks[1] if len(output_tasks) > 1 else None
            except FileNotFoundError:
                tqdm.write(f"B{index + 1} - File not found in cell: {row['subfolder_name']}/{row['file_name']}")
                continue 
        return df
                
    
    def get_tasks(self, output: str, tasks: list) -> list:
        """Extracts tasks from the AI model output.
        """
        output_tasks = []
        for task_group in tasks:
            for key, tasks in task_group.items():
                for task in tasks:
                    if task in output:
                        output_tasks.append(task)
        return output_tasks
        
    
    def tasks_string(self, task: list) -> str:
        """Nicely formats the list of tasks into a string for the AI model.
        """
        string = ''
        for i, t in enumerate(task):
            string += (f"{t}\n").replace('{', '').replace('}', '').replace("'","")
        return string
    
    
            
    def ai_chat(self, model_source: str, model: str, input: str) -> None:
        """
        Args:
            model_source (string): Model source. Either 'ollama' or 'OpenAI'.
            model (string): Model name.
            txt (string): The text to send to the model.

        Returns:
            string: The response from the model.
        """
        if model_source == 'ollama':
            client = Client(host='http://localhost:11434')
            response = client.chat(
                model=model,
                messages=[{'role': 'user', 'content': f'{input}/n'}]
            )
            return response.message.content.strip()

        elif model_source == 'OpenAI':
            # Get API key from config or .env file
            if not self.config['api_key']:
                try:
                    load_dotenv()
                    self.config['api_key'] = os.getenv('API_KEY')
                    if not self.config['api_key']:
                        raise ValueError("API key not found in .env file.")
                        
                except Exception as e:
                    print(f"\nError loading API key: {e}")
                    exit()

            client = OpenAI(api_key=self.config['api_key'])
            completion = client.chat.completions.create(
                model=model,
                messages=[{"role": "user",
                    "content": input,
                }])
            return completion.choices[0].message.content

        else:
            print(f"\nUnknown model source: {model_source}")
            exit()


    def save_to_csv(self, df: pd.DataFrame) -> None:
        """Saves the DataFrame to a CSV file."""
        df.to_csv(os.path.join(os.getcwd() + self.config['csv_dir'].format(datetime=strftime('%Y-%m-%d %H-%M'))), index=True, columns=['Name', 'origTask', 'newTask1', 'newTask2'], encoding='utf-8-sig')


    def save_to_xlsx(self, df: pd.DataFrame) -> None:
        #! Not in use, as comments are removed from the Excel file. 
        wb = load_workbook(os.path.join(os.getcwd() + self.config['file_dir']))
        ws = wb.active
        # write to the Excel file
        for index, row in df.iterrows():
            ws.cell(row=index + 2, column=5, value=row['newTask1'])
            ws.cell(row=index + 2, column=6, value=row['newTask2'])

        # Save the workbook
        wb.save(os.path.join(os.getcwd() + self.config['file_dir']))




# RUN 
if __name__ == "__main__":
    task = Task()
    task.run()