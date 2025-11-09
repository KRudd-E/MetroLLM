from src.utils import get_config, pureAPI_query, check_config
from src.retrieve import Retriever
from src.ai import AI_Assister

import os
import csv
import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook
from time import strftime


class Pipeline:
    def __init__(self):
        self.config = get_config()
        self.starting_datetime = strftime("%Y-%m-%d %H-%M-%S")
        self.ai_assist = AI_Assister(self.config)
        self.retriever = Retriever(self.config)
        self.csv_path = os.path.join(os.getcwd() + '/' + self.config['output_csv_dir'].replace('.csv', f'-{self.starting_datetime}.csv'))
        self.csv_initialized = False

    
    def run(self):

        #**** Initialise ****#
        pureAPI_query(self.config)
        check_config(self.config)
        print('')
        
        #**** Get Subdirectory list and Subdirectory-Names list ****#
        subdirs:        list = sorted([x[0] for x in os.walk(os.getcwd() + self.config['txt_grandparent_dir'])][1:])
        subdir_names:   list = sorted([x[1] for x in os.walk(os.getcwd() + self.config['txt_grandparent_dir']) if x[1] != []][0])
        
        #**** Filter subdirs and subdir_names based on subdir_ending_selection ****#
        if self.config['subdir_ending_selection']:
            subdirs:        list = [x for x in subdirs if x.endswith(self.config['subdir_ending_selection'])]
            subdir_names:   list = [x for x in subdir_names if x.endswith(self.config['subdir_ending_selection'])]
            
        #**** Use starting_subfolder_manager to adjust starting subfolder if specified ****#
        if self.config['beginning_dir']:
            subdir_names, subdirs = self.starting_subfolder_manager(self.config['beginning_dir'], subdir_names, subdirs)
        #**** Use ending_subfolder_manager to adjust starting subfolder if specified ****#
        if self.config['ending_dir']:
            subdir_names, subdirs = self.ending_subfolder_manager(self.config['ending_dir'], subdir_names, subdirs)
        
        #**** Create DataFrame and ID ***#
        df = pd.DataFrame(columns=['id', 'File Name', \
            'Folder Name', 'Name', 'Level', 'Country', \
                'Year', 'Sector(s)', 'Task(s)', 'Object Keyword(s)', \
                'Measurement extent', 'Measurement Tolerance', \
                'Surface Interaction', 'Measured Object Properties', \
                'Tools and Methods', 'Environment Properties', 'Task Operation', \
                'User', 'User branch location or group', \
                'User partners', 'Info source', 'System source', \
                'System type', 'Model', 'New/Old'])
        id: int = self.config['beginning_id']
        

        #**** Iterate over folders and text files ****#
        for subdir, subdir_name in tqdm(zip(subdirs, subdir_names), total=len(subdirs), desc="Processing Subdirectories", position=0, colour='blue'):
            
            text_files: list = sorted([f for f in os.listdir(subdir) if f.endswith('.txt')])
            
            for file_name in tqdm(text_files, total=len(text_files), desc=f"Processing '{subdir_name}'", position=1, colour='green', leave=False):
                
                #**** Read the case study text file ***#
                with open(os.path.join(subdir + '/' + file_name), 'r') as file:
                    case_study_text: str = file.read()
                
                #**** Write empty line for an invalid case study ****#
                if len(case_study_text) < self.config['min_case_study_length']:
                    tqdm.write(f"Case study {file_name} is too short - {len(case_study_text)} chars. Writing empty row.")
                    row: pd.Series = self.create_row(id, file_name, subdir_name, {})
                    self.save_to_excel(row, file_name, id)
                    id += 1
                    continue



                #**** Level & Name ****#
                if self.config['use_level_name']:
                    level_name_dict: dict = self.retriever.retrieve_multiple(
                        names=['level', 'name'],
                        options=None,
                        prompt=self.config['level_name_prompt'].format(txt=case_study_text)
                    )
                else:
                    level_name_dict = {'level': '', 'name': ''}
                
                #**** Country & Year ****#
                if self.config['use_country_year']:
                    country_year_dict: dict = self.retriever.retrieve_multiple(
                        names=['country', 'year'],
                        options=None,
                        prompt=self.config['country_year_prompt'].format(txt=case_study_text)
                    )
                else:
                    country_year_dict = {'country': '', 'year': ''}
                
                #**** Sector(s) ****#
                if self.config['use_sector']:
                    sector_dict: dict = self.retriever.retrieve_multiple(
                        names=['sector'],
                        options={
                            'sector': self.config['sector_list_v2']
                        },
                        prompt=self.config['sector_prompt'].format(
                            sector_list=self.config['sector_list_v2'],
                            txt=case_study_text),
                        tries=3,
                        expect_result=True
                    )
                else:
                    sector_dict: dict = {'sector': ''}
                
                #**** Task(s) ****#
                if self.config['use_task']:
                    #? Check for diverse case study : multiple parts to the case study
                    if not self.is_diverse_check(case_study_text, file_name):
                        
                        task_dict: dict = self.retriever.retrieve_multiple(
                            names=['task'],
                            options={
                                'task': self.config['task_list']
                            },
                            prompt=self.config['task_prompt'].format(
                                task_list=self.config['task_list'],
                                txt=case_study_text
                            ),
                            tries=3,
                            expect_result=True,
                        )
                    
                    else:
                        task_dict: dict = {'task': ['Diverse']}
                else:
                    task_dict: dict = {'task': ''}
                
                #**** Object Keyword(s) ****#
                if self.config['use_object_keywords']:
                    object_keywords_dict: dict = self.retriever.retrieve_multiple(
                        names=['object keyword'],
                        
                        options={'object keyword': self.config['object_keywords_list']},
                        
                        prompt=self.config['object_keywords_prompt'].format(
                            object_keywords_list=self.config['object_keywords_list'],
                            txt=case_study_text),
                        tries=4,
                        expect_result=True
                    )
                else:
                    object_keywords_dict: dict = {'object keyword': ''}
                
                #**** Measurement Metrics ****#
                if self.config['use_measurement_metrics']:
                    measurement_metrics_dict: dict = self.retriever.retrieve_multiple(
                        names=['measurement extent','measurement tolerance'],
                        
                        options={'measurement extent': self.config['measurement_extent_list_v2'],
                                    'measurement tolerance': self.config['measurement_tolerance_list_v1']},
                        
                        prompt=self.config['measurement_metrics_prompt'].format(
                            measurement_extent_list=self.config['measurement_extent_list_v2'],
                            measurement_tolerance_list=self.config['measurement_tolerance_list_v1'],
                            txt=case_study_text),
                    )
                    
                else:
                    measurement_metrics_dict: dict = {'measurement extent': '', 'measurement tolerance': ''}
                
                #**** Surface Interaction & Measured Object Properties ****#
                if self.config['use_surface_interaction_and_object_properties']:
                    surface_interaction_and_measured_object_properties_dict: dict = self.retriever.retrieve_multiple(
                        names=['surface interaction', 'measured object properties'],
                        
                        options={
                            'surface interaction': self.config['surface_interaction_list_v2'],
                            'measured object properties': self.config['measured_object_properties_list_v2']
                        },
                        
                        prompt=self.config['surface_interaction_and_measured_object_properties_prompt'].format(
                            surface_interaction_list=self.config['surface_interaction_list_v2'],
                            measured_object_properties_list=self.config['measured_object_properties_list_v2'],
                            txt=case_study_text)
                    )
                    surface_interaction_and_measured_object_properties_dict = self.remove_brackets_from_dict_vals(surface_interaction_and_measured_object_properties_dict)
                else:
                    surface_interaction_and_measured_object_properties_dict: dict = {'surface interaction': '', 'object properties': ''}
                
                #**** Tools and Methods ****#
                if self.config['use_tools_and_methods']:
                    tools_and_methods_dict: dict = self.retriever.retrieve_multiple(
                        names=['tools and methods'],
                        options={'tools and methods': self.config['tools_methods_list']},
                        
                        prompt=self.config['tools_methods_prompt'].format(
                            tools_methods_list=self.config['tools_methods_list'],
                            txt=case_study_text)
                    )
                else:
                    tools_and_methods_dict: dict = {'tools and methods': ''}
                
                #**** Environment and Task Operations & User Model ****#
                if self.config['use_environment_and_task_operations']:
                    environment_and_task_operations_dict: dict = self.retriever.retrieve_multiple(
                        names=['environment properties', 'task operation'],
                        options={
                            'environment properties': self.config['environment_list'],
                            'task operation': self.config['task_operation_list_v2']
                        },
                        prompt=self.config['environment_operation_prompt'].format(
                            environment_list=self.config['environment_list'],
                            task_operation_list=self.config['task_operation_list_v2'],
                            txt=case_study_text),
                        tries=3,
                        expect_result=True
                    )
                    environment_and_task_operations_dict = self.remove_brackets_from_dict_vals(environment_and_task_operations_dict)
                else:
                    environment_and_task_operations_dict: dict = {'environment properties': '', 'task operation': ''}
                
                #**** User Model ****#
                if self.config['use_user_model']:
                    user_model_dict: dict = self.retriever.retrieve_multiple(
                        names=['user', 'user branch location or group', 'user partners', 'model'],
                        options=None,
                        prompt=self.config['user_model_prompt'].format(
                            txt=case_study_text,
                            company=subdir_name.replace(' - UNPROCESSED', '')
                            ),
                        print_responses=False,
                    )
                else:
                    user_model_dict: dict = {'user': '', 'user branch location or group': '', 'user partners': '', 'model': ''}
                
                #**** New/Old Case Study ****#
                if self.config['if_ending']:
                    new_old_dict: dict = {'new/old': 'New' if subdir_name.endswith(self.config['if_ending_word']) else 'Old'}
                else:
                    new_old_dict: dict = {'new/old': ''}
                
                
                #**** Combine dictionaries ****#
                info_dict: dict = level_name_dict | country_year_dict | sector_dict | task_dict |\
                    object_keywords_dict | measurement_metrics_dict | surface_interaction_and_measured_object_properties_dict |\
                    tools_and_methods_dict | environment_and_task_operations_dict | user_model_dict | new_old_dict
                
                #**** Save info ****#
                row: pd.Series = self.create_row(id, file_name, subdir_name, info_dict)
                df = pd.concat([df, row.to_frame().T], ignore_index=True)
            
                self.save_to_excel(row, file_name, id)
                self.save_to_csv_iterative(row, id)
                id += 1
                
                
                                              
    def save_to_csv_iterative(self, row: pd.Series, id: int) -> None:
            """Saves the row to a CSV file iteratively, creating the file with headers on first call."""
            try:
                if not self.csv_initialized:
                    with open(self.csv_path, mode='w', newline='', encoding='utf-8') as f:
                        writer = csv.writer(f)
                        writer.writerow(row.index.tolist())
                    self.csv_initialized = True
                with open(self.csv_path, mode='a', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(row.tolist())
            except Exception as e:
                tqdm.write(f"Error saving to CSV during id {id}: {e}")
                exit(1)

    def create_row(self, id: int, file_name: str, subdir_name: str, info_vals: dict) -> pd.Series:
        """ Creates a row for the DataFrame with the given values. Prints to terminal if terminal_outputs is True.

        Args:
            id (int): ID of the row
            file_name (str): name of the file being processed
            subdir_name (str): name of the subdirectory containing the file
            info_vals (dict): dictionary containing the values to be added to the row

        Returns:
            pd.Series: A pandas Series object representing the row to be added to the DataFrame
        """
        row = pd.Series({
            'id': id,
            'File Name': file_name.replace('.txt', '.pdf'),
            'Folder Name': subdir_name,
            'Name': self.clean_value(info_vals.get('name', '')),
            'Level': self.clean_value(info_vals.get('level', '')),
            'Country': self.clean_value(info_vals.get('country', '')),
            'Year': self.clean_value(info_vals.get('year', '')),
            'Sector(s)': self.clean_value(info_vals.get('sector', '')),
            'Task(s)': self.clean_value(info_vals.get('task', '')),
            'Object Keyword(s)': self.clean_value(info_vals.get('object keyword', '')),
            'Measurement extent': self.clean_value(info_vals.get('measurement extent', '')),
            'Measurement Tolerance': self.clean_value(info_vals.get('measurement tolerance', '')),
            'Surface Interaction': self.clean_value(info_vals.get('surface interaction', '')),
            'Measured Object Properties': self.clean_value(info_vals.get('measured object properties', '')),
            'Tools and Methods': self.clean_value(info_vals.get('tools and methods', '')),
            'Environment Properties': self.clean_value(info_vals.get('environment properties', '')),
            'Task Operation': self.clean_value(info_vals.get('task operation', '')),
            'User': self.clean_value(info_vals.get('user', '')),
            'User branch location or group': self.clean_value(info_vals.get('user branch location or group', '')),
            'User partners': self.clean_value(info_vals.get('user partners', '')),
            'Model': self.clean_value(info_vals.get('model', '')),
            'New/Old': self.clean_value(info_vals.get('new/old', '')),
            # 'Info source': '',
            # 'System source': '',
            # 'System type': '',
        })
        
        if self.config['terminal_outputs']:
            tqdm.write(f"{file_name} - {subdir_name} - {id} - {row.to_dict()}")
        
        return row


    def save_to_excel(self, row: pd.Series , file_name: str, id: int) -> None:
        """ Saves the row to an Excel file, creating a new sheet and first line if it's the first iteration.
        """
        try:         
            wb = load_workbook(f'{os.getcwd()}/{self.config["output_xlsx_dir"]}')
            
            # Create new sheet on first iteration
            if id == self.config['beginning_id']:
                ws = wb.create_sheet(title=self.starting_datetime)
            
            ws = wb[self.starting_datetime]

            # Write header if it's the first row
            if id == self.config['beginning_id']:
                ws.append(row.index.tolist())

            # Append and save row
            ws.append(row.tolist())
            
            # add hyperlink to name
            if self.config['use_level_name']:
                ws.cell(row=ws.max_row, column=4).hyperlink = f"{os.getcwd()}/data/1-original/applicationsDB/{row['Folder Name']}/{file_name[:-4]}.pdf" #type: ignore
                ws.cell(row=ws.max_row, column=4).style = 'Hyperlink'
            
            wb.save(f'{os.getcwd()}/{self.config["output_xlsx_dir"]}')
        except Exception as e:
            tqdm.write(f"Error saving to Excel during {file_name}: {e}")
            exit(1)

    @staticmethod
    def clean_value(val: str | list) -> str:
        """ Cleans the value to be added to the DataFrame.
        """
        if isinstance(val, list):
            if len(val) == 0:
                return ''
            elif len(val) == 1:
                return str(val[0])
            else:
                return ' | '.join(str(v) for v in val)
        return val if val is not None else ''

    @staticmethod
    def starting_subfolder_manager(starting_subfolder: str, subdir_names: list, subdirs: list) -> tuple:
        """ Adjusts subdirs and subdir_names to start from the specified starting_subfolder.
        """
        if starting_subfolder not in subdir_names:
            raise ValueError(f"Starting subfolder '{starting_subfolder}' not found in subdirectories list. Check config.yaml and data.")
        sbf_idx = subdir_names.index(starting_subfolder)
        subdirs = subdirs[sbf_idx:]
        subdir_names = subdir_names[sbf_idx:]    
        return subdir_names, subdirs
    
    @staticmethod
    def ending_subfolder_manager(ending_subfolder: str, subdir_names: list, subdirs: list) -> tuple:
        """ Adjusts subdirs and subdir_names to start from the specified starting_subfolder.
        """
        if ending_subfolder not in subdir_names:
            raise ValueError(f"Ending subfolder '{ending_subfolder}' not found in subdir_names\n Check config.yaml and data.")
        sbf_idx = subdir_names.index(ending_subfolder) + 1
        subdirs = subdirs[:sbf_idx]
        subdir_names = subdir_names[:sbf_idx]    
        return subdir_names, subdirs


    
    def is_diverse_check(self, case_study_text, file_name) -> bool:
        input_text = self.config['diverse_check_prompt'].format(
            txt=case_study_text
        )
        for i in range(2):
            response = self.ai_assist.ai_chat(
                model_source=self.config['model_source'],
                model=self.config['model'],
                input=input_text
            )
            # get last line of response
            response = response.strip().split('\n')[-1].strip()
            
            # Check if the response contains 'Yes' or 'No' (along with other text)
            if 'Yes' in response or 'No' in response:
                # Return True if 'Yes' is found, False if 'No' is found
                return True if 'Yes' in response else False
        tqdm.write(f"\nDiversity check failed for case study: {file_name}")  # Log the failure
        return False 

    def remove_brackets_from_dict_vals(self, input_dict: dict) -> dict:
        for key, val in input_dict.items():
            if isinstance(val, str):
                # Remove content in brackets from the value
                input_dict[key] = val.split('(')[0].strip()
            elif isinstance(val, list):
                # Process lists to remove brackets from each item
                input_dict[key] = [item.split('(')[0].strip() for item in val]
            elif isinstance(val, dict):
                # Recursively process nested dictionaries
                input_dict[key] = self.remove_brackets_from_dict_vals(val)

        return input_dict