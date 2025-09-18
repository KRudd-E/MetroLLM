import os
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm
from src.utils.utils import tc_app_query, tc_def_query, tc_comp_query
from src.utils.format_txt import format_txt

class TextClass:
    def __init__(self, config, src):
        self.config = config
        if src.lower() == 'applications'  : self.applicationsDB_Gen(config['applicationsDB'])
        if src.lower() == 'definitions'   : self.definitionsDB_Gen(config['definitionsDB'])
        if src.lower() == 'companies'     : self.companiesDB_Gen(config['companiesDB'])

    def applicationsDB_Gen(self, config):
        tc_app_query()

        #** Get data from Excel doc **#
        excel_path = os.path.join(os.getcwd() + config['xlsx_dir'])
        pure_xlsx_df = pd.read_excel(
            excel_path,
            sheet_name=config['sheet_name'],
            index_col=None
        )
        pure_xlsx_df.rename(columns={'File': 'Name'}, inplace=True)
        pure_xlsx_df = pure_xlsx_df[['Name', 'Date', 'Sector', 'Task', 'User', 'Location', 'Level']]
        
        wb = load_workbook(excel_path)
        ws = wb[config['sheet_name']] if config['sheet_name'] in wb.sheetnames else wb.active
        assert ws is not None, "Worksheet not found in the workbook."
        
        db_base = os.path.join(os.getcwd() + config['db_dir'])

        # Add new columns for text data
        pure_xlsx_df['id'] = None
        pure_xlsx_df['Subdir'] = None
        pure_xlsx_df['File'] = None
        pure_xlsx_df['Text'] = None
        pure_xlsx_df['Text_len'] = None

        #** Process each row in the worksheet **#
        for i, row in tqdm(enumerate(ws.iter_rows(min_row=2, min_col=2, max_col=2), start=0), 
                           total=ws.max_row - 1, desc="Processing rows"):
            cell = row[0]
            try:
                pdf_path = cell.hyperlink.target.replace('%20', ' ') #type: ignore
                file_name = (pdf_path.split('\\')[-1])[:-4] + '.txt'
                subfolder_name = pdf_path.split('\\')[-2]
                subfolder_path = os.path.join(db_base + subfolder_name)
                #tqdm.write(f"Processing file: {file_name} in subfolder: {subfolder_name}") #debugging output

                if not os.path.exists(subfolder_path):
                    print(f"Missing subfolder: {subfolder_path}")
                    continue

                with open(os.path.join(subfolder_path, file_name), 'r', encoding='utf-8') as file:
                    txt = file.read()
                    txt = format_txt(txt)

                    pure_xlsx_df.at[i, 'id'] = i + 1
                    pure_xlsx_df.at[i, 'Text'] = txt
                    pure_xlsx_df.at[i, 'Text_len'] = len(txt)
                    pure_xlsx_df.at[i, 'Subdir'] = subfolder_name
                    pure_xlsx_df.at[i, 'File'] = file_name

            except (AttributeError, FileNotFoundError) as e:
                print(f"{cell.coordinate} - {e}")
                continue

        #** Select & reorder columns **#
        df = pure_xlsx_df[['id', 'Name', 'Subdir', 'File', 'Task', 'Sector', 'User', 'Location', 'Level', 'Text_len', 'Text']]
        df['id'].astype(int)

        if 'rows_to_remove' in config and isinstance(config['rows_to_remove'], list):
            df = df[~df['id'].isin(config['rows_to_remove'])]


        #** Map tasks **#
        expanded_tasks = []
        task_map = config.get('task_map', {})

        for i, row in df.iterrows():
            task = str(row['Task']).strip()
            
            if task in task_map:
                mapped_tasks = task_map[task]
                row['Task'] = mapped_tasks if isinstance(mapped_tasks, list) else [mapped_tasks]
                expanded_tasks.append(row)
            else:
                row['Task'] = [task]
                expanded_tasks.append(row)


        #** Save to CSV **#
        df = pd.DataFrame(expanded_tasks).reset_index(drop=True)
        df['id'] = range(1, len(df) + 1)
        df.to_csv(os.path.join(os.getcwd() + config['output_dir']), index=False, encoding='utf-8', quoting="", escapechar='\\') # type: ignore
        
    
    
    
    
    #! Placeholders    
    def definitionsDB_Gen(self, config):
        tc_def_query()
        print("definitionsDB_Gen not implemented yet.")
        pass
    
    
    def companiesDB_Gen(self, config):
        tc_comp_query()
        print("companiesDB_Gen not implemented yet.")
        pass

